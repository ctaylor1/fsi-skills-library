#!/usr/bin/env python3
"""Regenerate catalog/skills-catalog.{json,csv} from the source build-plan workbook.

The workbook is the authoritative plan and is maintained outside this repository.
Pass its path with --source, or set FSI_BUILD_PLAN_XLSX.

Usage:
    python tools/build_catalog_from_xlsx.py --source "<path-to>.xlsx"
"""
from __future__ import annotations
import argparse, csv, json, os, sys
from pathlib import Path

DEFAULT_SOURCE = os.environ.get(
    "FSI_BUILD_PLAN_XLSX",
    r"C:/Users/Taylor/OneDrive/02 - Work/Projects/2026-07-Amazon Quick FSI Skills/"
    r"amazon-quick-fsi-skills-2026-catalog-and-build-plan-v2.xlsx",
)
REPO = Path(__file__).resolve().parents[1]

# Skill-type taxonomy (six types). Primary assignment is derived from the build archetype;
# a small curated override set handles skills whose content clearly indicates another type.
DEFAULT_TYPE_BY_ARCHETYPE = {
    "Explain & summarize": "Guidance or domain-expertise skills",
    "Analyze & review": "Analysis and evaluation skills",
    "Model & calculate": "Artifact-creation skills",
    "Draft & package": "Artifact-creation skills",
    "Reconcile & validate": "Analysis and evaluation skills",
    "Monitor & alert": "System-interaction or operational skills",
    "Investigate & casework": "Analysis and evaluation skills",
    "Domain workflow": "Guidance or domain-expertise skills",
    "Orchestrate & resolve": "Workflow or orchestration skills",
}
SKILL_TYPE_OVERRIDES = {
    # Utility: small, reusable, technical parsers/transformers/tooling
    "iso-20022-message-interpreter": "Utility skills",
    "financials-normalizer": "Utility skills",
    "fsi-skill-authoring-assistant": "Utility skills",
    # System-interaction / operational: run against and coordinate live operational systems
    "security-alert-triage-assistant": "System-interaction or operational skills",
    "cyber-incident-response-coordinator": "System-interaction or operational skills",
    "omnichannel-case-orchestrator": "System-interaction or operational skills",
}


def skill_type_for(name: str, archetype: str) -> str:
    return SKILL_TYPE_OVERRIDES.get(name) or DEFAULT_TYPE_BY_ARCHETYPE.get(archetype, "Guidance or domain-expertise skills")


def rows(ws, header_row: int):
    hdr = [c.value for c in ws[header_row]]
    out = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in r):
            continue
        out.append({hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]})
    return out


def build(source: str) -> list[dict]:
    import openpyxl  # imported here so --help works without the dependency

    wb = openpyxl.load_workbook(source, data_only=True)
    catalog = rows(wb["Skill Catalog"], 4)
    matrix = rows(wb["Build Matrix"], 4)
    midx = {(m["Category"], m["Use Case"]): m for m in matrix if m.get("Use Case")}

    def g(d, k):
        return (d.get(k) or "").strip() if d else ""

    skills = []
    for c in catalog:
        if not c.get("Use Case"):
            continue
        m = midx.get((c["Category"], c["Use Case"]), {})
        _name = c["Use Case"].strip()
        _arch = g(m, "Build Archetype")
        skills.append({
            "name": _name,
            "category": c["Category"].strip(),
            "description": g(c, "Description"),
            "primary_user": g(c, "Primary User"),
            "status": g(c, "Status"),
            "skill_type": skill_type_for(_name, _arch),
            "archetype": _arch,
            "agent_pattern": g(m, "Agent Pattern"),
            "risk_tier": g(m, "Risk Tier"),
            "action_mode": g(m, "Default Action Mode"),
            "human_approval": g(m, "Human Approval"),
            "delivery_wave": g(m, "Delivery Wave"),
            "tooling_data": g(m, "Minimum Tooling / Data"),
            "validation_focus": g(m, "Primary Validation Focus"),
            "scheduled_agent": g(m, "Scheduled-Agent Candidate"),
        })
    return skills


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--source", default=DEFAULT_SOURCE)
    args = ap.parse_args()

    if not Path(args.source).exists():
        print(f"ERROR: source workbook not found: {args.source}", file=sys.stderr)
        return 2

    skills = build(args.source)
    out_dir = REPO / "catalog"
    out_dir.mkdir(exist_ok=True)
    (out_dir / "skills-catalog.json").write_text(
        json.dumps(
            {"portfolio": "Amazon Quick Desktop — 2026 FSI Skill Portfolio v2",
             "count": len(skills), "skills": skills},
            indent=2, ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    with (out_dir / "skills-catalog.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(skills[0].keys()))
        w.writeheader()
        w.writerows(skills)
    print(f"Wrote {len(skills)} skills to catalog/skills-catalog.json and .csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
